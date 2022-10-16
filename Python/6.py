from openpyxl import load_workbook
import string
AZUpper = string.ascii_uppercase

workbook = load_workbook(filename="Book.xlsx")
workbook.sheetnames


sheet = workbook.active
n = 11

for i in range(len(AZUpper)) :
  if AZUpper[i] == "K" :

    break
  n -= 1
  sheet[AZUpper[i] + "1"] = n



workbook.save(filename=("Book.xlsx"))