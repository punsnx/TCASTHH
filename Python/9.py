import sys
from openpyxl import load_workbook
import string
import random
AZUpper = string.ascii_uppercase

workbook = load_workbook(filename="minesweeper.xlsx")
workbook.sheetnames

sheet = workbook.active
for c in range(len(AZUpper)) :
  for cc in range(1,27,1) :
    sheet[AZUpper[c] + str(cc)] = ""

for b in range(10) :
  sheet[AZUpper[random.randint(0,25)] + str(random.randint(1,26))] = "Bomb"

workbook.save(filename=("minesweeper.xlsx"))


def fb(l,n) :
  xy = l + str(n)
  if sheet[xy].value == "Bomb" :
    print(xy + " is a Bomb! [Finished]")
    sheet[xy].value = "! " + xy + " Bomb"
    workbook.save(filename=("minesweeper.xlsx"))
    sys.exit()
  elif int(n) <= 26 and int(n) != 0:
    print(xy + " is not a Bomb! [NOOB!]")
    sheet[xy].value = xy
  else :
    print("error!")
  workbook.save(filename=("minesweeper.xlsx"))
  pass
i=0
while i < 1 :
  p = str(input("Input A1 - Z26 : "))
  if p == "" :
    print("please fill")
  elif len(p) < 2 or len(p) > 3 :
    print("error!")
  elif len(p) == 2 :
    fb(p[0], p[1])
  elif len(p) == 3 :
    fb(p[0], str(p[1]) + str(p[2]))